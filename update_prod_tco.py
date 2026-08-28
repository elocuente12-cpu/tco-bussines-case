"""
Reconstruye el TCO del ambiente PRODUCTIVO combinando:
  (1) Los 12 servidores PLANEADOS (aun no en IaC, pero se desplegaran) tomados
      del assessment original (analysis_TCO backup). Algunos iran en ASG (EC2
      names consecutivos) y otros standalone.
  (2) Los componentes YA DESPLEGADOS en la IaC (pro): 2 RDS, FSx, DMS.

Reglas confirmadas con el usuario:
  - El "ASG app1" desplegado (USAEA1PWBWES21/22) ES el servidor planeado
    PAHWPAPTUI03 -> se FUSIONAN en una sola fila (no duplicar 21-22).
  - Se INCLUYE el AMI Builder (m5.large, temporal).
  - Los servidores que van en ASG NO llevan EBS adicional (col M vacia): su
    almacenamiento es el FSx compartido.
  - Costos: se CONSERVAN los del assessment donde existen; se CALCULAN con
    precios reales AWS us-east-1 los faltantes (pahqpapapp02, pahwpapbof01) y
    la columna O (EBS anual) de todas las filas EC2/RDS/DMS/FSx.

Precios: AWS Price List us-east-1, shared tenancy, RI Standard No-Upfront,
730 h/mes, anualizado x12. Nota: RDS SQL SE license-included NO tiene RI 3yr
No-Upfront (celdas RI3 = NA).

Total: 17 filas (12 planeados + RDS produccion + RDS procesos + FSx + AMI + DMS).
"""
import openpyxl
from copy import copy

FILEPATH = '/Users/javier.sepulveda/projects/experian/tco-bussines-case/Ambiente Productivo/Lift-and-Shift - 18 Servers/analysis_TCO.xlsx'

H = 730
M = 12
def yr(hr):
    return None if hr is None else round(hr * H * M, 2)

EBS_GP3 = 0.08
FSX_SSD_SINGLE = 0.130
FSX_TP_SINGLE = 2.20
DMS_STORAGE = 0.115
def ebs_yr(gb, mult=1):
    return None if gb is None else round(gb * EBS_GP3 * M * mult, 2)

# Precios $/hr (para filas sin costo en assessment)
r7a = {'od': 0.12208, 'ri1': 0.09632, 'ri3': 0.08051}
r7a_infra = {'od': 0.07608, 'ri1': 0.05032, 'ri3': 0.03451}
c5a_lin = {'od': 0.077, 'ri1': 0.049, 'ri3': 0.033}  # Linux, sin licencia
m5 = {'od': 0.188, 'ri1': 0.152, 'ri3': 0.133}
m5_infra = {'od': 0.096, 'ri1': 0.060, 'ri3': 0.041}
rds4xl = {'od': 12.16, 'ri1': 11.4912}
rds4xl_nolic = {'od': 5.952}
rds2xl = {'od': 3.04, 'ri1': 2.8728}
rds2xl_nolic = {'od': 1.488}
dms_hr = 0.476

# ---------------------------------------------------------------------------
# Filas (22 columnas A..V). Los 12 planeados conservan costos del assessment
# (P..V) salvo pahqpapapp02 y pahwpapbof01 que se calculan. La columna O (EBS)
# se calcula para todos. ASG => ebs_add=None.
# ---------------------------------------------------------------------------
def r(host, ec2, cpus, ram, ostype, osname, inst, cores, awsram, ebs_root, ebs_add,
      ebs_cost, od_total, lic, od_excl, ri1_total, ri1_excl, ri3_total, ri3_excl):
    return dict(host=host, ec2=ec2, env='PROD', cpus=cpus, ram=ram, ostype=ostype, osname=osname,
                rec=inst, dep=inst, cores=cores, awsram=awsram, ebs_root=ebs_root, ebs_add=ebs_add,
                region='US East (N. Virginia)', ebs_cost=ebs_cost, od_total=od_total, lic=lic,
                od_excl=od_excl, ri1_total=ri1_total, ri1_excl=ri1_excl, ri3_total=ri3_total, ri3_excl=ri3_excl)

rows = [
    # === 12 servidores planeados (assessment) ===
    # standalone RHEL (sin EC2 name asignado)
    r('PAHWPAPSTS01', None, 4, 16, 'RHEL', 'Red Hat Enterprise Linux 8 (64-bit)', 'r5a.large', 2, 16, 520, None,
      ebs_yr(520), 1332, 0, 1244, 964, 876, 771, 683),
    r('PAHWPAPSTE01', None, 4, 16, 'RHEL', 'Red Hat Enterprise Linux 8 (64-bit)', 'r5a.large', 2, 16, 210, None,
      ebs_yr(210), 1332, 0, 1244, 964, 876, 771, 683),
    # ASG (EC2 names consecutivos) -> EBS adicional VACIO (storage = FSx)
    r('PAHWPAPSRC03', 'USAEA1PWBWES23\nUSAEA1PWBWES24', 4, 16, 'Windows', 'Microsoft Windows Server 2022 (64-bit)', 'r5a.large', 2, 16, 120, None,
      ebs_yr(120), 1858, 806, 990, 1490, 622, 1276, 569),
    r('PAHWPAPWFC01', 'USAEA1PWBWES25\nUSAEA1PWBWES26', 4, 8, 'Windows', 'Microsoft Windows Server 2019 (64-bit)', 'r7a.medium', 1, 8, 80, None,
      ebs_yr(80), 1113, 403, 666, 888, 441, 749, 302),
    # standalone Windows
    r('PAHWPWFINT01', 'USAEA1PFSWES09', 4, 8, 'Windows', 'Microsoft Windows Server 2019 (64-bit)', 'r7a.medium', 1, 8, 80, 300,
      ebs_yr(80), 1113, 403, 666, 888, 441, 749, 302),
    r('PAHWPWSAPI01', 'USAEA1PWBWES35', 4, 16, 'Windows', 'Microsoft Windows Server 2022 (64-bit)', 'r5a.large', 2, 16, 120, 90,
      ebs_yr(120), 1858, 806, 990, 1490, 622, 1276, 569),
    # ASG
    r('PAHWPAPWSS03', 'USAEA1PWBWES27\nUSAEA1PWBWES28', 4, 8, 'Windows', 'Microsoft Windows Server 2022 (64-bit)', 'r7a.medium', 1, 8, 120, None,
      ebs_yr(120), 1113, 403, 666, 888, 441, 749, 302),
    # standalone
    r('PAHWPAPWEF02', 'USAEA1PAPWES34', 4, 8, 'Windows', 'Microsoft Windows Server 2022 (64-bit)', 'r7a.medium', 1, 8, 120, 90,
      ebs_yr(120), 1113, 403, 666, 888, 441, 749, 302),
    # ASG (21-22) == ASG app1 desplegado. FUSIONADO en esta fila.
    r('PAHWPAPTUI03', 'USAEA1PWBWES21\nUSAEA1PWBWES22', 4, 8, 'Windows', 'Microsoft Windows Server 2022 (64-bit) [ASG app1 desplegado]', 'r7a.medium', 1, 8, 120, None,
      ebs_yr(120), 1113, 403, 666, 888, 441, 749, 302),
    # standalone
    r('PAHWPCHECS01', 'USAEA1PFSWES08', 4, 16, 'Windows', 'Microsoft Windows Server 2022 (64-bit)', 'r5a.large', 2, 16, 120, 90,
      ebs_yr(120), 1858, 806, 990, 1490, 622, 1276, 569),
    # pahqpapapp02 - sin costos en assessment -> CALCULADO (r7a.medium Windows)
    r('pahqpapapp02', 'USAEA1PAPWES33', 4, 8, 'Windows', 'Windows Server 2019 Standard', 'r7a.medium', 4, 8, 280, None,
      ebs_yr(280), yr(r7a['od']), yr(r7a['od']-r7a_infra['od']), yr(r7a_infra['od']),
      yr(r7a['ri1']), yr(r7a_infra['ri1']), yr(r7a['ri3']), yr(r7a_infra['ri3'])),
    # pahwpapbof01 - sin costos en assessment -> CALCULADO (c5a.large Linux)
    r('pahwpapbof01', None, 4, 2, 'Linux', 'Ubuntu 22.04.3 LTS', 'c5a.large', 2, 4, 210, None,
      ebs_yr(210), yr(c5a_lin['od']), 0, yr(c5a_lin['od']),
      yr(c5a_lin['ri1']), yr(c5a_lin['ri1']), yr(c5a_lin['ri3']), yr(c5a_lin['ri3'])),

    # === Componentes YA DESPLEGADOS en la IaC ===
    r('RDS produccion (multiaz1)', None, 16, 128, 'SQL Server - Standard',
      'SQL Server Standard Edition 15.x (Multi-AZ, license-included)', 'db.r6i.4xlarge', 16, 128, 1700, None,
      ebs_yr(1700, 2), yr(rds4xl['od']), yr(rds4xl['od']-rds4xl_nolic['od']), yr(rds4xl_nolic['od']),
      yr(rds4xl['ri1']), None, 'NA', 'NA'),
    r('RDS procesos (standalone1)', None, 8, 64, 'SQL Server - Standard',
      'SQL Server Standard Edition 15.x (Single-AZ, license-included)', 'db.r6i.2xlarge', 8, 64, 2900, None,
      ebs_yr(2900, 1), yr(rds2xl['od']), yr(rds2xl['od']-rds2xl_nolic['od']), yr(rds2xl_nolic['od']),
      yr(rds2xl['ri1']), None, 'NA', 'NA'),
    r('FSx Windows (intelisrcpa-prd)', None, None, None, 'Windows (FSx)',
      'FSx Windows File Server Single-AZ SSD (64 MBps, backup 30d)', 'FSx SINGLE_AZ_1', None, None, 100, None,
      round(100*FSX_SSD_SINGLE*M, 2), round(64*FSX_TP_SINGLE*M, 2), 0, round(64*FSX_TP_SINGLE*M, 2),
      None, None, None, None),
    r('AMI Builder (temporal)', None, 2, 8, 'Windows', 'Microsoft Windows Server 2025 (AMI Builder - temporal)',
      'm5.large', 2, 8, 100, None,
      ebs_yr(100), yr(m5['od']), yr(m5['od']-m5_infra['od']), yr(m5_infra['od']),
      yr(m5['ri1']), yr(m5_infra['ri1']), yr(m5['ri3']), yr(m5_infra['ri3'])),
    r('DMS replication (poc)', None, 8, 16, 'DMS', 'DMS c5.2xlarge Single-AZ (engine 3.6.1)',
      'dms.c5.2xlarge', 8, 16, 50, None,
      round(50*DMS_STORAGE*M, 2), yr(dms_hr), 0, yr(dms_hr), None, None, None, None),
]

wb = openpyxl.load_workbook(FILEPATH)
ws = wb['Shared Tenancy Analysis']

# Capturar estilo de una fila de datos existente (fila 2)
data_style = []
for c in range(1, 23):
    cell = ws.cell(2, c)
    data_style.append({'font': copy(cell.font), 'alignment': copy(cell.alignment),
                       'border': copy(cell.border), 'number_format': cell.number_format})

# Limpiar filas de datos
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.value = None

def vals(s):
    return [s['host'], s['ec2'], s['env'], s['cpus'], s['ram'], s['ostype'], s['osname'],
            s['rec'], s['dep'], s['cores'], s['awsram'], s['ebs_root'], s['ebs_add'], s['region'],
            s['ebs_cost'], s['od_total'], s['lic'], s['od_excl'],
            s['ri1_total'], s['ri1_excl'], s['ri3_total'], s['ri3_excl']]

for idx, s in enumerate(rows, start=2):
    for c, v in enumerate(vals(s), start=1):
        cell = ws.cell(row=idx, column=c)
        cell.value = v
        st = data_style[c-1]
        cell.font = st['font']; cell.alignment = st['alignment']
        cell.border = st['border']; cell.number_format = st['number_format']

# Eliminar filas sobrantes
last = 1 + len(rows)
if ws.max_row > last:
    ws.delete_rows(last+1, ws.max_row - last)

# Des-ocultar filas: el archivo original del assessment tenia filas marcadas
# como hidden que ocultaban recursos (RDS/FSx/DMS). Aseguramos que todas las
# filas de datos sean visibles.
for _r, _dim in list(ws.row_dimensions.items()):
    if _dim.hidden:
        _dim.hidden = False

wb.save(FILEPATH)

print(f"TCO productivo reconstruido con {len(rows)} filas.")
asg = [s for s in rows if s['ec2'] and '\n' in str(s['ec2'])]
print(f"  Servidores planeados: 12 | Desplegados (RDS/FSx/AMI/DMS): 5 | En ASG (sin EBS add): {len(asg)}")
print("\nDetalle:")
for s in rows:
    ec2 = (str(s['ec2']).replace('\n','/') if s['ec2'] else '')
    print(f"  {s['host']:<28} {ec2:<26} {str(s['dep']):<16} root={s['ebs_root']} add={s['ebs_add']}")
