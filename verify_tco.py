"""
Verifica el TCO de Ambientes Intermedios contra el estado esperado de la IaC.
Muestra la tabla resultante y valida coherencia (dimensiones y costos anualizados
recalculados a partir de las columnas mensuales, dado que las formulas =PRODUCT
solo se evaluan al abrir Excel).
"""
import openpyxl

FILEPATH = '/Users/javier.sepulveda/projects/experian/tco-bussines-case/Ambientes intermedios/Assessment-2026-03-04-1706/analysis_TCO.xlsx'

# Estado esperado segun IaC desplegada (repo intelisrcpa)
# host -> (env, instance, cores, ram, ebs_root, ebs_add)
EXPECTED = {
    'PAHQTAPIIS01':          ('UAT', 'm5a.xlarge',   4, 16, 130, None),
    'PAHQTFSDSC01':          ('UAT', 'c5a.xlarge',   4,  8, 100, None),
    'PAHQTWFINT01':          ('UAT', 'c5a.xlarge',   4,  8,  80, None),
    'PAHQTAPSTE01':          ('UAT', 'c5a.xlarge',   4,  8, 210, None),
    'PAHQTAPSTS01':          ('UAT', 'm5.xlarge',    4, 16, 520, None),
    'PAHQTAPBOF01':          ('QA',  'm5.large',     2,  8,  80, None),
    'SSRP':                  ('DEV', 'm5.xlarge',    4, 16, 100, None),
    'PAHQDAPIIS01':          ('DEV', 'm5a.large',    2,  8, 100, None),
    'PAHQDWFINT01':          ('DEV', 'c5a.xlarge',   4,  8, 100, None),
    'PAHQDAPWSS01':          ('QA',  'm5.large',     2,  8, 100, 100),
    'PAHQTWFINT02':          ('QA',  'c5a.xlarge',   4,  8,  80, None),
    'RDS SQL Server (UAT)':  ('UAT', 'db.r5.large',  2, 16, 200, None),
    'RDS SQL Server (DEV)':  ('DEV', 'db.m6i.xlarge',4, 16, 500, None),
    'RDS SQL Server (QA)':   ('QA',  'db.r5.large',  2, 16, 200, None),
}

wb = openpyxl.load_workbook(FILEPATH, data_only=True)
ws = wb['Shared Tenancy Analysis']

print("=== TCO Ambientes Intermedios - Shared Tenancy Analysis ===\n")
hdr = f"{'Host':<22}{'Env':<5}{'Instance':<14}{'Cor':>4}{'RAM':>5}{'Root':>6}{'Add':>5}  {'OD/mo':>9}{'RI1/mo':>9}{'RI3/mo':>9}"
print(hdr)
print("-" * len(hdr))

issues = []
for r in range(2, ws.max_row + 1):
    host = ws.cell(r, 1).value
    if not host:
        continue
    env   = ws.cell(r, 3).value
    cores = ws.cell(r, 10).value
    ram   = ws.cell(r, 11).value
    inst  = ws.cell(r, 9).value
    root  = ws.cell(r, 12).value
    add   = ws.cell(r, 13).value
    od    = ws.cell(r, 19).value
    ri1   = ws.cell(r, 21).value
    ri3   = ws.cell(r, 23).value
    add_s = '-' if add is None else add
    print(f"{host:<22}{str(env):<5}{str(inst):<14}{str(cores):>4}{str(ram):>5}{str(root):>6}{str(add_s):>5}  "
          f"{str(od):>9}{str(ri1):>9}{str(ri3):>9}")

    # Validacion contra IaC
    if host in EXPECTED:
        exp_env, exp_inst, exp_cores, exp_ram, exp_root, exp_add = EXPECTED[host]
        if env != exp_env:      issues.append(f"{host}: env {env} != IaC {exp_env}")
        if inst != exp_inst:    issues.append(f"{host}: instance {inst} != IaC {exp_inst}")
        if cores != exp_cores:  issues.append(f"{host}: cores {cores} != IaC {exp_cores}")
        if ram != exp_ram:      issues.append(f"{host}: RAM {ram} != IaC {exp_ram}")
        if root != exp_root:    issues.append(f"{host}: EBS root {root} != IaC {exp_root}")
        if add != exp_add:      issues.append(f"{host}: EBS add {add} != IaC {exp_add}")

print(f"\nTotal registros: {ws.max_row - 1}")
print("\n=== Validacion contra IaC desplegada ===")
if not issues:
    print("  OK - Todas las filas coinciden con la IaC (env, instancia, cores, RAM, EBS root/adicional).")
else:
    for i in issues:
        print(f"  DIFF: {i}")

# Totales por ambiente (On-Demand mensual y RI-3yr mensual)
print("\n=== Totales mensuales por ambiente (On-Demand / RI 3yr) ===")
tot = {}
for r in range(2, ws.max_row + 1):
    env = ws.cell(r, 3).value
    od  = ws.cell(r, 19).value
    ri3 = ws.cell(r, 23).value
    if env is None:
        continue
    od = od if isinstance(od, (int, float)) else 0
    ri3 = ri3 if isinstance(ri3, (int, float)) else 0
    t = tot.setdefault(env, [0, 0])
    t[0] += od
    t[1] += ri3
gtot = [0, 0]
for env in ['UAT', 'DEV', 'QA']:
    if env in tot:
        od, ri3 = tot[env]
        gtot[0] += od
        gtot[1] += ri3
        print(f"  {env:<5} On-Demand ${od:>10,.2f}/mes  |  RI-3yr ${ri3:>10,.2f}/mes")
print(f"  {'TOTAL':<5} On-Demand ${gtot[0]:>10,.2f}/mes  |  RI-3yr ${gtot[1]:>10,.2f}/mes")
print(f"  {'':<5} On-Demand ${gtot[0]*12:>10,.2f}/anio |  RI-3yr ${gtot[1]*12:>10,.2f}/anio")
