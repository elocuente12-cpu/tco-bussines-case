"""
Actualiza el TCO de Ambientes Intermedios (dev/qa/stg) para reflejar el estado
REAL de la IaC desplegada (repo intelisrcpa: dev, qa, stg).

Enfoque NO destructivo: edita solo las celdas que cambian y preserva las
formulas existentes (columnas P, T, V, X = =PRODUCT(x,12)).

Fuente de precios: AWS Price List (us-east-1), shared tenancy, RI Standard
No-Upfront. Metodologia validada: reproduce exactamente los valores previos
del Excel para las filas sin cambios (730 h/mes).

Precios $/hora us-east-1:
  c5a.xlarge RHEL: OD 0.212 / RI-1yr 0.155 / RI-3yr 0.124
  m5.xlarge  RHEL: OD 0.250 / RI-1yr 0.179 / RI-3yr 0.141
EBS root gp3: 0.08 USD/GB-mes (EC2)
"""
import openpyxl

FILEPATH = '/Users/javier.sepulveda/projects/experian/tco-bussines-case/Ambientes intermedios/Assessment-2026-03-04-1706/analysis_TCO.xlsx'
HOURS = 730

wb = openpyxl.load_workbook(FILEPATH)
ws = wb['Shared Tenancy Analysis']

changes = []

def setc(row, col, value, note):
    # Asignacion via .value para que el borrado (None) persista correctamente
    # incluso sobre celdas que contenian formulas.
    old = ws.cell(row, col).value
    ws.cell(row=row, column=col).value = value
    changes.append((f"{openpyxl.utils.get_column_letter(col)}{row}", old, value, note))

# ---------------------------------------------------------------------------
# R5 - PAHQTAPSTE01 (UAT, Linux/RHEL): c5a.large -> c5a.xlarge (IaC: c5a.xlarge)
#   vCPU 2->4, RAM 4->8 GB. EBS root 210 GB (sin cambio).
# ---------------------------------------------------------------------------
setc(5, 4, 4, "CPUs 2->4 (c5a.xlarge)")            # D
setc(5, 5, 8, "RAM 4->8 GB")                        # E
setc(5, 8, 'c5a.xlarge', "AWS Instance Recommended") # H
setc(5, 9, 'c5a.xlarge', "AWS Instance Deploy")      # I
setc(5, 10, 4, "AWS Total Cores 2->4")               # J
setc(5, 11, 8, "AWS RAM 4->8 GB")                    # K
setc(5, 19, round(0.212 * HOURS, 2), "Monthly On-Demand (c5a.xlarge RHEL)")  # S = 154.76
setc(5, 21, round(0.155 * HOURS, 2), "Monthly 1yr RI")  # U = 113.15
setc(5, 23, round(0.124 * HOURS, 2), "Monthly 3yr RI")  # W = 90.52
# O (EBS root) = 210 * 0.08 = 16.8 -> ya correcto, no se toca

# ---------------------------------------------------------------------------
# R6 - PAHQTAPSTS01 (UAT, Linux/RHEL): m5.large -> m5.xlarge (IaC: m5.xlarge)
#   vCPU 2->4, RAM 8->16 GB. EBS root 520 GB (sin cambio).
# ---------------------------------------------------------------------------
setc(6, 4, 4, "CPUs 2->4 (m5.xlarge)")             # D
setc(6, 5, 16, "RAM 8->16 GB")                      # E
setc(6, 8, 'm5.xlarge', "AWS Instance Recommended") # H
setc(6, 9, 'm5.xlarge', "AWS Instance Deploy")      # I
setc(6, 10, 4, "AWS Total Cores 2->4")              # J
setc(6, 11, 16, "AWS RAM 8->16 GB")                 # K
setc(6, 19, round(0.250 * HOURS, 2), "Monthly On-Demand (m5.xlarge RHEL)")  # S = 182.5
setc(6, 21, round(0.179 * HOURS, 2), "Monthly 1yr RI")  # U = 130.67
setc(6, 23, round(0.141 * HOURS, 2), "Monthly 3yr RI")  # W = 102.93

# ---------------------------------------------------------------------------
# R7 - PAHQTAPBOF01: reubicado de UAT a QA en la IaC (env tst).
#   Environment UAT->QA, EC2 name USAEA1UAPLES03 -> USAEA1TAPLES01.
#   Instancia m5.large Linux (Ubuntu), root 80 GB: sin cambios de dimension
#   ni costo (m5.large Linux = 0.096/hr = 70.08/mes).
# ---------------------------------------------------------------------------
setc(7, 2, 'USAEA1TAPLES01', "EC2 Name (QA naming)")  # B
setc(7, 3, 'QA', "Environment UAT->QA (reubicado en IaC)")  # C

# ---------------------------------------------------------------------------
# R12 - PAHQTWFINT02 (QA): limpiar EBS adicional fantasma.
#   La IaC de QA NO define disco adicional para esta instancia; el Excel tenia
#   Q12=8 (Monthly Additional EBS) y R12=formula. Se eliminan para reflejar la
#   realidad (solo root de 80 GB).
# ---------------------------------------------------------------------------
setc(12, 17, None, "Monthly Additional EBS Cost -> vacio (no hay disco adicional)")  # Q
setc(12, 18, None, "Annualized Additional EBS Cost -> vacio")                        # R

wb.save(FILEPATH)

print("=== Cambios aplicados al TCO (Shared Tenancy Analysis) ===")
for coord, old, new, note in changes:
    print(f"  {coord}: {old!r} -> {new!r}   [{note}]")
print(f"\nTotal de celdas modificadas: {len(changes)}")
print("Nota: columnas P, T, V, X se recalculan automaticamente (formulas =PRODUCT(x,12)).")
