import openpyxl

filepath = '/Users/javier.sepulveda/projects/experian/tco-bussines-case/Ambiente Productivo/Lift-and-Shift - 18 Servers/analysis.xlsx'
wb = openpyxl.load_workbook(filepath, data_only=True)

print("=== SHEET NAMES ===")
print(wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n\n=== SHEET: {sheet_name} (rows={ws.max_row}, cols={ws.max_column}) ===")
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50), values_only=True):
        filtered = [str(c) for c in row if c is not None]
        if filtered:
            print(" | ".join(filtered))
