"""
Verifica el TCO productivo actualizado contra el inventario real de la IaC (pro).
Muestra la tabla resultante y valida que cada componente esperado este presente
con su instancia y storage correctos.
"""
import openpyxl

FILEPATH = '/Users/javier.sepulveda/projects/experian/tco-bussines-case/Ambiente Productivo/Lift-and-Shift - 18 Servers/analysis_TCO.xlsx'

# Estado esperado segun IaC (pro). instance -> (cores, ram, ebs_root)
EXPECTED = {
    'USAEA1PWBWES21': ('r7a.medium', 1, 8, 120),
    'USAEA1PWBWES22': ('r7a.medium', 1, 8, 120),
    'RDS produccion (multiaz1)':  ('db.r6i.4xlarge', 16, 128, 1700),
    'RDS procesos (standalone1)': ('db.r6i.2xlarge', 8, 64, 2900),
    'FSx Windows (intelisrcpa-prd)': ('FSx SINGLE_AZ_1', None, None, 100),
    'AMI Builder (temporal)': ('m5.large', 2, 8, 100),
    'DMS replication (poc)': ('dms.c5.2xlarge', 8, 16, 50),
}

wb = openpyxl.load_workbook(FILEPATH, data_only=True)
ws = wb['Shared Tenancy Analysis']

print("=== TCO Productivo (segun IaC desplegada) - Shared Tenancy Analysis ===\n")
hdr = f"{'Componente':<30}{'EC2 Name':<18}{'Instance':<16}{'Cor':>4}{'RAM':>5}{'Stor':>6}  {'EBS/yr':>9}{'OD/yr':>11}{'Lic/yr':>10}{'RI1/yr':>11}{'RI3/yr':>9}"
print(hdr); print('-'*len(hdr))

issues = []
tot = {'ebs':0,'od':0,'lic':0,'ri1':0}
for r in range(2, ws.max_row+1):
    host = ws.cell(r,1).value
    if not host: continue
    ec2  = ws.cell(r,2).value or ''
    inst = ws.cell(r,9).value
    cores= ws.cell(r,10).value
    ram  = ws.cell(r,11).value
    stor = ws.cell(r,12).value
    ebsc = ws.cell(r,15).value
    od   = ws.cell(r,16).value
    lic  = ws.cell(r,17).value
    ri1  = ws.cell(r,19).value
    ri3  = ws.cell(r,21).value
    def n(x): return x if isinstance(x,(int,float)) else 0
    tot['ebs']+=n(ebsc); tot['od']+=n(od); tot['lic']+=n(lic); tot['ri1']+=n(ri1)
    print(f"{str(host):<30}{str(ec2)[:16]:<18}{str(inst):<16}{str(cores):>4}{str(ram):>5}{str(stor):>6}  "
          f"{str(ebsc):>9}{str(od):>11}{str(lic):>10}{str(ri1):>11}{str(ri3):>9}")
    # validar
    key = ec2 if ec2 in EXPECTED else host
    if key in EXPECTED:
        e_inst,e_cores,e_ram,e_stor = EXPECTED[key]
        if inst!=e_inst:  issues.append(f"{key}: instance {inst} != IaC {e_inst}")
        if cores!=e_cores:issues.append(f"{key}: cores {cores} != IaC {e_cores}")
        if ram!=e_ram:    issues.append(f"{key}: RAM {ram} != IaC {e_ram}")
        if stor!=e_stor:  issues.append(f"{key}: storage {stor} != IaC {e_stor}")
    else:
        issues.append(f"Fila no esperada: {host} / {ec2}")

# cobertura
present = set()
for r in range(2, ws.max_row+1):
    ec2=ws.cell(r,2).value or ''; host=ws.cell(r,1).value
    present.add(ec2 if ec2 in EXPECTED else host)
for k in EXPECTED:
    if k not in present: issues.append(f"FALTA en Excel: {k}")

print(f"\nTotal componentes: {ws.max_row-1}")
print("\n=== Validacion contra IaC (pro) ===")
if not issues:
    print("  OK - Todos los componentes coinciden con la IaC (instancia, cores, RAM, storage).")
else:
    for i in issues: print(f"  DIFF: {i}")

print("\n=== Totales anuales ===")
print(f"  EBS/Storage:        ${tot['ebs']:>12,.2f}")
print(f"  On-Demand compute:  ${tot['od']:>12,.2f}  (EC2 + RDS + DMS + FSx throughput)")
print(f"  License-only:       ${tot['lic']:>12,.2f}")
print(f"  1yr NURI compute:   ${tot['ri1']:>12,.2f}  (parcial: RDS SE 3yr NU no aplica)")
print(f"  TOTAL OD + storage: ${tot['od']+tot['ebs']:>12,.2f}/anio")
