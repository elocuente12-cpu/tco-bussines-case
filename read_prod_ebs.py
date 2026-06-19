import openpyxl

filepath = '/Users/javier.sepulveda/projects/experian/tco-bussines-case/Ambiente Productivo/Lift-and-Shift - 18 Servers/analysis.xlsx'
wb = openpyxl.load_workbook(filepath, data_only=True)

# Read Block-EBS-Cost-Optimized to understand storage per server
ws = wb['Block-EBS-Cost-Optimized']
print("=== Block-EBS-Cost-Optimized ===")
print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")

# Get headers
headers = [cell.value for cell in ws[1]]
print(f"\nHeaders:")
for i, h in enumerate(headers):
    print(f"  Col {i}: {h}")

# Get all data rows
print(f"\n{'Row':<4} {'Capacity Total':<15} {'Capacity Used':<15} {'Boot Vol Size':<15}")
print("-" * 60)
for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
    cap_total = row[18] if len(row) > 18 else ''
    cap_used = row[19] if len(row) > 19 else ''
    boot_size = row[12] if len(row) > 12 else ''
    print(f"{idx:<4} {str(cap_total):<15} {str(cap_used):<15} {str(boot_size):<15}")
