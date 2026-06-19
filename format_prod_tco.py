import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from copy import copy

# Paths
prod_filepath = '/Users/javier.sepulveda/projects/experian/tco-bussines-case/Ambiente Productivo/Lift-and-Shift - 18 Servers/analysis.xlsx'

# Load production workbook
wb_prod = openpyxl.load_workbook(prod_filepath, data_only=True)

# ============================================================
# 1. Extract server data from all 3 pricing sheets
# ============================================================
def extract_servers_from_sheet(ws):
    servers = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        # Only process rows that have a valid server_id (starts with "server-")
        if row[0] is None:
            continue
        server_id = str(row[0]).strip()
        if not server_id.startswith('server-'):
            continue
        servers.append({
            'server_id': row[0],
            'server_name': row[1],
            'host_name': row[2],
            'cluster_name': row[3],
            'hypervisor': row[4],
            'environment': row[5],
            'application': row[6],
            'server_type': row[7],
            'num_cpus': row[8],
            'cores_per_cpu': row[9],
            'total_cores': row[10],
            'ram': row[11],
            'os_type': row[12],
            'os_name': row[13],
            'peak_cpu': row[14],
            'avg_cpu': row[15],
            'peak_ram': row[16],
            'avg_ram': row[17],
            'uptime': row[18],
            'ec2_recommended': row[19],
            'ec2_cores': row[20],
            'ec2_ram': row[21],
            'region': row[22],
            'annualized_total': row[23],
            'annualized_network': row[24],
            'annualized_license': row[25],
            'annualized_ec2_excl_license': row[26],
        })
    return servers

servers_od = extract_servers_from_sheet(wb_prod['Shared Tenancy - On-Demand'])
servers_1yr = extract_servers_from_sheet(wb_prod['Shared Tenancy - 1yr NU'])
servers_3yr = extract_servers_from_sheet(wb_prod['Shared Tenancy - 3yr NU'])

print(f"Servers On-Demand: {len(servers_od)}")
print(f"Servers 1yr NU: {len(servers_1yr)}")
print(f"Servers 3yr NU: {len(servers_3yr)}")

# Extract extra servers at end of on-demand sheet
ws_od = wb_prod['Shared Tenancy - On-Demand']
extra_servers = []
found_extra = False
for row in ws_od.iter_rows(min_row=2, max_row=ws_od.max_row, values_only=True):
    if row[0] and str(row[0]).strip() == 'Server-Hostname':
        found_extra = True
        continue
    if found_extra and row[0] is not None:
        extra_servers.append({
            'server_name': row[0],
            'ip': row[1],
            'vcpu': row[2],
            'memory': row[3],
            'storage_gb': row[4],
            'environment': row[5],
            'os_name_simple': row[6],
            'os_version': row[7],
            'app_description': row[8],
        })

print(f"Extra servers (manual): {len(extra_servers)}")
for s in extra_servers:
    print(f"  {s['server_name']} - {s['vcpu']} vCPU - {s['memory']} GB RAM - {s['storage_gb']} GB - {s['os_name_simple']}")

# ============================================================
# 2. Map EBS volumes from Block-EBS-Cost-Optimized
# ============================================================
ws_ebs = wb_prod['Block-EBS-Cost-Optimized']
ebs_volumes = []
for row in ws_ebs.iter_rows(min_row=2, max_row=ws_ebs.max_row, values_only=True):
    ebs_volumes.append({
        'volume_name': row[0],
        'capacity_total': row[18],
        'capacity_used': row[19],
        'boot_vol_size': row[12],
    })

print(f"\nTotal EBS volumes: {len(ebs_volumes)}")

# Build disk pairs per server
disk_pairs = []
current_pair = []
for vol in ebs_volumes:
    vol_name = str(vol['volume_name'])
    if vol_name.endswith('-disk-0'):
        if current_pair:
            disk_pairs.append(current_pair)
        current_pair = [vol]
    elif vol_name.endswith('-disk-1'):
        current_pair.append(vol)
    else:
        current_pair.append(vol)
if current_pair:
    disk_pairs.append(current_pair)

print(f"Disk pairs (servers with EBS): {len(disk_pairs)}")

# Map to the 18 servers from assessment
server_names_ordered = [s['server_name'] for s in servers_od]
ebs_per_server = {}
for i, server_name in enumerate(server_names_ordered):
    if i < len(disk_pairs):
        pair = disk_pairs[i]
        root_size = pair[0]['capacity_total']
        additional = pair[1]['capacity_total'] if len(pair) > 1 else None
        ebs_per_server[server_name] = {'root': root_size, 'additional': additional}
    else:
        ebs_per_server[server_name] = {'root': None, 'additional': None}

# ============================================================
# 3. Create new workbook matching intermedios format
# ============================================================
wb_new = openpyxl.Workbook()
wb_new.remove(wb_new.active)

# --- Sheet 1: Read Me ---
ws_readme = wb_new.create_sheet("Read Me")
ws_readme['A1'] = "This file contains an export of AWS Transform Assessment results for Production Environment (InteliSrcPA)."
ws_readme['A3'] = "Sheet"
ws_readme['B3'] = "Description"
ws_readme['A4'] = "Glossary"
ws_readme['B4'] = "Column-by-column definitions and descriptions for all data fields."
ws_readme['A5'] = "Shared Tenancy Analysis"
ws_readme['B5'] = "Lists servers that can run on shared hardware infrastructure in AWS."

# --- Sheet 2: Glossary ---
ws_glossary = wb_new.create_sheet("Glossary")
glossary_items = [
    ("Attribute Name", "Description"),
    ("Shared Tenancy Analysis", ""),
    ("Host Name Onpremise", "Name assigned to the server on-premise"),
    ("EC2 Name", "Name for the EC2 instance in AWS"),
    ("Environment", "Production/Development/Test environment designation"),
    ("Number of CPUs", "Count of virtual processors"),
    ("RAM (GB)", "Total memory capacity in gigabytes"),
    ("Operation System Type", "Category of operating system"),
    ("Operation System Name", "Specific OS version/distribution"),
    ("AWS Instance Recommended", "Recommended AWS instance type by assessment tool"),
    ("AWS Instance Deploy", "Instance type to be deployed"),
    ("AWS Total Cores", "Total cores in recommended EC2 instance"),
    ("AWS RAM (GB)", "Memory in recommended instance in GB"),
    ("EBS Size (GB)", "Root volume EBS storage size in GB"),
    ("EBS Size (GB) Additional", "Additional EBS volumes in GB"),
    ("AWS Region", "Target AWS geographic region"),
    ("Annualized 1 Yr EBS Cost", "Yearly EBS storage cost"),
    ("Annualized On-Demand Total EC2 - RDS Cost", "Yearly total cost using on-demand pricing"),
    ("Annualized License Only Cost", "Yearly licensing costs"),
    ("Annualized On-Demand EC2 Cost Excl. License Cost", "Yearly compute costs at on-demand rates"),
    ("Annualized 1 Yr NURI Total EC2 - RDS Cost", "Total yearly cost with 1-year RI no upfront"),
    ("Annualized 1 Yr NURI EC2 Cost, Excl. License Costs", "Compute-only yearly cost with 1-year RI"),
    ("Annualized 3 Yr NURI Total EC2 Cost", "Total yearly cost with 3-year RI no upfront"),
    ("Annualized 3 Yr NURI EC2 Cost Excl. License Cost", "Compute-only yearly cost with 3-year RI"),
]
for idx, (attr, desc) in enumerate(glossary_items, start=1):
    ws_glossary.cell(row=idx, column=1, value=attr)
    ws_glossary.cell(row=idx, column=2, value=desc)

# --- Sheet 3: Shared Tenancy Analysis ---
ws_main = wb_new.create_sheet("Shared Tenancy Analysis")

headers = [
    "Host Name Onpremise",
    "EC2 Name",
    "Enviroment",
    "Number of CPUs",
    "RAM (GB)",
    "Operation System Type",
    "Operation System Name",
    "AWS Instance Recommended",
    "AWS Instance Deploy",
    "AWS Total Cores",
    "AWS RAM (GB)",
    "EBS Size (GB)",
    "EBS Size (GB)",
    "AWS Region",
    "Annualized 1 Yr EBS Cost",
    "Annualized On-Demand Total EC2 - RDS Cost",
    "Annualized License Only Cost",
    "Annualized On-Demand EC2 Cost Excl. License Cost",
    "Annualized 1 Yr NURI Total EC2 - RDS Cost",
    "Annualized 1 Yr NURI EC2 Cost, Excl. License Costs",
    "Annualized 3 Yr NURI Total EC2 Cost",
    "Annualized 3 Yr NURI EC2 Cost Excl. License Cost",
]

header_font = Font(bold=True)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
for col_idx, header in enumerate(headers, start=1):
    cell = ws_main.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

# Write server data
row_idx = 2
for i, srv_od in enumerate(servers_od):
    name = srv_od['server_name']
    srv_1yr = servers_1yr[i] if i < len(servers_1yr) else {}
    srv_3yr = servers_3yr[i] if i < len(servers_3yr) else {}
    ebs = ebs_per_server.get(name, {'root': None, 'additional': None})
    
    os_type = srv_od.get('os_type', '')
    if os_type == 'Win':
        os_type_full = 'Windows'
    elif os_type == 'RHEL':
        os_type_full = 'RHEL'
    else:
        os_type_full = os_type
    
    ws_main.cell(row=row_idx, column=1, value=name)
    ws_main.cell(row=row_idx, column=2, value='')
    ws_main.cell(row=row_idx, column=3, value='PROD')
    ws_main.cell(row=row_idx, column=4, value=srv_od.get('total_cores'))
    ws_main.cell(row=row_idx, column=5, value=srv_od.get('ram'))
    ws_main.cell(row=row_idx, column=6, value=os_type_full)
    ws_main.cell(row=row_idx, column=7, value=srv_od.get('os_name'))
    ws_main.cell(row=row_idx, column=8, value=srv_od.get('ec2_recommended'))
    ws_main.cell(row=row_idx, column=9, value=srv_od.get('ec2_recommended'))
    ws_main.cell(row=row_idx, column=10, value=srv_od.get('ec2_cores'))
    ws_main.cell(row=row_idx, column=11, value=srv_od.get('ec2_ram'))
    ws_main.cell(row=row_idx, column=12, value=ebs.get('root'))
    ws_main.cell(row=row_idx, column=13, value=ebs.get('additional'))
    ws_main.cell(row=row_idx, column=14, value='US East (N. Virginia)')
    ws_main.cell(row=row_idx, column=15, value=None)
    ws_main.cell(row=row_idx, column=16, value=srv_od.get('annualized_total'))
    ws_main.cell(row=row_idx, column=17, value=srv_od.get('annualized_license'))
    ws_main.cell(row=row_idx, column=18, value=srv_od.get('annualized_ec2_excl_license'))
    ws_main.cell(row=row_idx, column=19, value=srv_1yr.get('annualized_total') if srv_1yr else None)
    ws_main.cell(row=row_idx, column=20, value=srv_1yr.get('annualized_ec2_excl_license') if srv_1yr else None)
    ws_main.cell(row=row_idx, column=21, value=srv_3yr.get('annualized_total') if srv_3yr else None)
    ws_main.cell(row=row_idx, column=22, value=srv_3yr.get('annualized_ec2_excl_license') if srv_3yr else None)
    
    row_idx += 1

# Add the 3 extra servers
for extra in extra_servers:
    os_type_full = 'Windows' if 'Windows' in str(extra.get('os_name_simple', '')) else 'Linux'
    
    ws_main.cell(row=row_idx, column=1, value=extra['server_name'])
    ws_main.cell(row=row_idx, column=2, value='')
    ws_main.cell(row=row_idx, column=3, value='PROD')
    ws_main.cell(row=row_idx, column=4, value=extra.get('vcpu'))
    ws_main.cell(row=row_idx, column=5, value=extra.get('memory'))
    ws_main.cell(row=row_idx, column=6, value=os_type_full)
    ws_main.cell(row=row_idx, column=7, value=extra.get('os_version'))
    ws_main.cell(row=row_idx, column=8, value='')
    ws_main.cell(row=row_idx, column=9, value='')
    ws_main.cell(row=row_idx, column=10, value=extra.get('vcpu'))
    ws_main.cell(row=row_idx, column=11, value=extra.get('memory'))
    ws_main.cell(row=row_idx, column=12, value=extra.get('storage_gb'))
    ws_main.cell(row=row_idx, column=13, value=None)
    ws_main.cell(row=row_idx, column=14, value='US East (N. Virginia)')
    
    row_idx += 1

# Auto-adjust column widths
for col in ws_main.columns:
    max_length = 0
    column_letter = col[0].column_letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = min(max_length + 2, 40)
    ws_main.column_dimensions[column_letter].width = adjusted_width

# Save
output_path = '/Users/javier.sepulveda/projects/experian/tco-bussines-case/Ambiente Productivo/Lift-and-Shift - 18 Servers/analysis_TCO.xlsx'
wb_new.save(output_path)
print(f"\n✅ Archivo de produccion reformateado y guardado en:\n   {output_path}")
print(f"   Total servidores: {row_idx - 2} ({len(servers_od)} del assessment + {len(extra_servers)} manuales)")
